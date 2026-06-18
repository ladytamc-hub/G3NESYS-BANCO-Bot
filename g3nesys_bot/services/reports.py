from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..constants import DEFAULT_SETTINGS
from ..database import Database
from .callers import caller_ranking


MAGENTA = "C2185B"
DARK_MAGENTA = "55113A"
GOLD = "F4C542"
PALE_MAGENTA = "FCEAF4"
PALE_GOLD = "FFF7D6"
WHITE = "FFFFFF"
TEXT = "29232A"
LIGHT_BORDER = "E7D5DF"
CURRENCY_FORMAT = '#,##0 "plata";[Red](#,##0 "plata");-'
PERCENT_FORMAT = '0.0"%"'
DATE_FORMAT = "yyyy-mm-dd hh:mm"


def _generated_at() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _parse_date(value):
    if not value or isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return str(value)


def _member_name(guild, user_id: int | None) -> str:
    if not user_id:
        return ""
    member = guild.get_member(int(user_id)) if guild is not None else None
    return member.display_name if member is not None else "Usuario no disponible"


def _id(value) -> str:
    return str(value) if value not in (None, "") else ""


def _safe_sheet_name(name: str) -> str:
    return name[:31]


def _table_name(name: str) -> str:
    return "Tbl" + "".join(character for character in name.title() if character.isalnum())[:220]


def _style_title(sheet, title: str, last_column: int, subtitle: str) -> None:
    end = get_column_letter(max(last_column, 1))
    sheet.merge_cells(f"A1:{end}1")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor=DARK_MAGENTA)
    title_cell.font = Font(color=WHITE, bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(f"A2:{end}2")
    subtitle_cell = sheet["A2"]
    subtitle_cell.value = subtitle
    subtitle_cell.fill = PatternFill("solid", fgColor=PALE_GOLD)
    subtitle_cell.font = Font(color=TEXT, italic=True, size=9)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 20


def _add_detail_sheet(
    workbook: Workbook,
    *,
    name: str,
    title: str,
    subtitle: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    currency_headers: set[str] | None = None,
    percent_headers: set[str] | None = None,
    date_headers: set[str] | None = None,
) :
    sheet = workbook.create_sheet(_safe_sheet_name(name))
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, title, len(headers), subtitle)
    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = PatternFill("solid", fgColor=MAGENTA)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30

    normalized_rows = [list(row) for row in rows]
    if normalized_rows:
        for row in normalized_rows:
            sheet.append([_parse_date(value) for value in row])
        end_row = header_row + len(normalized_rows)
        table = Table(
            displayName=_table_name(name),
            ref=f"A{header_row}:{get_column_letter(len(headers))}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.cell(header_row + 1, 1, "Sin datos para este servidor.")
        end_row = header_row + 1

    currency_headers = currency_headers or set()
    percent_headers = percent_headers or set()
    date_headers = date_headers or set()
    thin = Side(style="thin", color=LIGHT_BORDER)
    for column, header in enumerate(headers, start=1):
        letter = get_column_letter(column)
        values = [str(header)] + [str(row[column - 1] or "") for row in normalized_rows]
        longest = max((len(value) for value in values), default=len(header))
        sheet.column_dimensions[letter].width = min(max(longest + 2, 11), 38)
        for row_number in range(header_row + 1, end_row + 1):
            cell = sheet.cell(row_number, column)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="right" if header in currency_headers | percent_headers else "left",
                vertical="top",
                wrap_text=longest > 28,
            )
            if header in currency_headers:
                cell.number_format = CURRENCY_FORMAT
            elif header in percent_headers:
                cell.number_format = PERCENT_FORMAT
            elif header in date_headers and isinstance(cell.value, datetime):
                cell.number_format = DATE_FORMAT
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{end_row}"
    return sheet


def _new_workbook(title: str, subtitle: str) -> tuple[Workbook, object]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, title, 12, subtitle)
    return workbook, sheet


def _save_workbook(workbook: Workbook, filename: str) -> Path:
    reports_dir = Path("data/reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    path = reports_dir / filename
    workbook.save(path)
    return path


def _populate_summary(
    sheet,
    *,
    metrics: Sequence[tuple[str, object, bool]],
    caller_rows: Sequence[tuple[str, int]],
    movement_rows: Sequence[tuple[str, int]],
    caller_chart_title: str = "Top callers por reputacion",
    movement_chart_title: str = "Movimientos por tipo",
) -> None:
    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["D"].width = 27
    sheet.column_dimensions["E"].width = 20
    sheet["A4"] = "Indicadores principales"
    sheet["A4"].font = Font(bold=True, color=WHITE, size=11)
    sheet["A4"].fill = PatternFill("solid", fgColor=MAGENTA)
    sheet.merge_cells("A4:E4")
    for index, (label, value, is_currency) in enumerate(metrics):
        row = 6 + index // 2 * 2
        column = 1 if index % 2 == 0 else 4
        label_cell = sheet.cell(row, column, label)
        value_cell = sheet.cell(row, column + 1, value)
        label_cell.fill = PatternFill("solid", fgColor=PALE_MAGENTA)
        label_cell.font = Font(bold=True, color=DARK_MAGENTA)
        value_cell.fill = PatternFill("solid", fgColor=PALE_GOLD)
        value_cell.font = Font(bold=True, color=TEXT, size=12)
        value_cell.alignment = Alignment(horizontal="right")
        if is_currency:
            value_cell.number_format = CURRENCY_FORMAT
        for cell in (label_cell, value_cell):
            cell.border = Border(
                left=Side(style="thin", color=LIGHT_BORDER),
                right=Side(style="thin", color=LIGHT_BORDER),
                top=Side(style="thin", color=LIGHT_BORDER),
                bottom=Side(style="thin", color=LIGHT_BORDER),
            )

    if caller_rows:
        sheet["P3"] = "Caller"
        sheet["Q3"] = "Puntos"
        for row_number, (name, score) in enumerate(caller_rows[:10], start=4):
            sheet.cell(row_number, 16, name)
            sheet.cell(row_number, 17, score)
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = caller_chart_title
        chart.y_axis.title = "Caller"
        chart.x_axis.title = "Puntos"
        data = Reference(sheet, min_col=17, min_row=3, max_row=3 + min(len(caller_rows), 10))
        categories = Reference(sheet, min_col=16, min_row=4, max_row=3 + min(len(caller_rows), 10))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 13
        chart.legend = None
        sheet.add_chart(chart, "G4")

    if movement_rows:
        start = 18
        sheet.cell(start, 16, "Tipo")
        sheet.cell(start, 17, "Monto")
        for row_number, (movement_type, amount) in enumerate(movement_rows[:10], start=start + 1):
            sheet.cell(row_number, 16, movement_type)
            sheet.cell(row_number, 17, amount)
        chart = BarChart()
        chart.type = "col"
        chart.style = 12
        chart.title = movement_chart_title
        chart.y_axis.title = "Plata"
        chart.x_axis.title = "Tipo"
        chart.y_axis.numFmt = '#,##0 "plata"'
        data = Reference(
            sheet,
            min_col=17,
            min_row=start,
            max_row=start + min(len(movement_rows), 10),
        )
        categories = Reference(
            sheet,
            min_col=16,
            min_row=start + 1,
            max_row=start + min(len(movement_rows), 10),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 13
        chart.legend = None
        sheet.add_chart(chart, "G20")

    sheet.column_dimensions["P"].hidden = True
    sheet.column_dimensions["Q"].hidden = True
    sheet.freeze_panes = "A4"


def _fetch_user_stats(db: Database, guild_id: int):
    return db.fetch_all(
        """
        WITH params(gid) AS (VALUES (?)),
        users AS (
            SELECT user_id FROM accounts, params WHERE guild_id = gid
            UNION SELECT user_id FROM fines, params WHERE guild_id = gid
            UNION SELECT user_id FROM withdrawals, params WHERE guild_id = gid
            UNION SELECT caller_id FROM activities, params WHERE guild_id = gid
            UNION SELECT ap.user_id FROM activity_participants ap
                JOIN activities ac ON ac.id = ap.activity_id, params WHERE ac.guild_id = gid
            UNION SELECT aa.usuario_id FROM asistencia_actividades aa
                JOIN activities ac ON ac.id = aa.actividad_id, params WHERE ac.guild_id = gid
            UNION SELECT pp.user_id FROM payout_participants pp
                JOIN payouts p ON p.id = pp.payout_id, params WHERE p.guild_id = gid
            UNION SELECT user_id FROM callers, params WHERE guild_id = gid
            UNION SELECT user_id FROM movements, params WHERE guild_id = gid AND user_id IS NOT NULL
            UNION SELECT counterparty_id FROM movements, params
                WHERE guild_id = gid AND counterparty_id IS NOT NULL
        )
        SELECT
            u.user_id,
            COALESCE(a.available, 0) AS available,
            COALESCE(a.retained, 0) AS retained,
            COALESCE(a.seized, 0) AS seized,
            COALESCE(a.available + a.retained + a.seized, 0) AS total_balance,
            a.updated_at AS account_updated_at,
            (SELECT COUNT(*) FROM fines f, params WHERE f.guild_id = gid AND f.user_id = u.user_id) AS fines_total,
            (SELECT COUNT(*) FROM fines f, params WHERE f.guild_id = gid AND f.user_id = u.user_id AND f.status = 'Pendiente') AS fines_pending,
            COALESCE((SELECT SUM(amount) FROM fines f, params WHERE f.guild_id = gid AND f.user_id = u.user_id), 0) AS fines_amount,
            (SELECT COUNT(*) FROM activity_participants ap JOIN activities ac ON ac.id = ap.activity_id, params WHERE ac.guild_id = gid AND ap.user_id = u.user_id) AS activities_joined,
            (SELECT COUNT(*) FROM asistencia_actividades aa JOIN activities ac ON ac.id = aa.actividad_id, params WHERE ac.guild_id = gid AND aa.usuario_id = u.user_id AND aa.estado = 'Confirmado') AS attendance_confirmed,
            (SELECT COUNT(*) FROM asistencia_actividades aa JOIN activities ac ON ac.id = aa.actividad_id, params WHERE ac.guild_id = gid AND aa.usuario_id = u.user_id AND aa.estado = 'Ausente') AS attendance_absent,
            (SELECT COUNT(*) FROM asistencia_actividades aa JOIN activities ac ON ac.id = aa.actividad_id, params WHERE ac.guild_id = gid AND aa.usuario_id = u.user_id AND aa.estado = 'Justificado') AS attendance_justified,
            (SELECT COUNT(*) FROM payout_participants pp JOIN payouts p ON p.id = pp.payout_id, params WHERE p.guild_id = gid AND pp.user_id = u.user_id) AS payouts_received,
            COALESCE((SELECT SUM(pp.amount) FROM payout_participants pp JOIN payouts p ON p.id = pp.payout_id, params WHERE p.guild_id = gid AND pp.user_id = u.user_id), 0) AS payouts_amount,
            (SELECT COUNT(*) FROM withdrawals w, params WHERE w.guild_id = gid AND w.user_id = u.user_id) AS withdrawals_total,
            COALESCE((SELECT SUM(amount_requested) FROM withdrawals w, params WHERE w.guild_id = gid AND w.user_id = u.user_id), 0) AS withdrawals_requested,
            COALESCE((SELECT SUM(amount_liquidated) FROM withdrawals w, params WHERE w.guild_id = gid AND w.user_id = u.user_id), 0) AS withdrawals_liquidated,
            (SELECT COUNT(*) FROM movements m, params WHERE m.guild_id = gid AND m.type = 'TRANSFERENCIA' AND m.user_id = u.user_id) AS transfers_sent,
            COALESCE((SELECT SUM(amount) FROM movements m, params WHERE m.guild_id = gid AND m.type = 'TRANSFERENCIA' AND m.user_id = u.user_id), 0) AS transfers_sent_amount,
            (SELECT COUNT(*) FROM movements m, params WHERE m.guild_id = gid AND m.type = 'TRANSFERENCIA' AND m.counterparty_id = u.user_id) AS transfers_received,
            COALESCE((SELECT SUM(amount) FROM movements m, params WHERE m.guild_id = gid AND m.type = 'TRANSFERENCIA' AND m.counterparty_id = u.user_id), 0) AS transfers_received_amount,
            EXISTS(SELECT 1 FROM callers c, params WHERE c.guild_id = gid AND c.user_id = u.user_id) AS is_caller,
            EXISTS(SELECT 1 FROM caller_penalties cp, params WHERE cp.guild_id = gid AND cp.user_id = u.user_id AND cp.active = 1) AS caller_penalized
        FROM users u
        LEFT JOIN accounts a ON a.guild_id = (SELECT gid FROM params) AND a.user_id = u.user_id
        ORDER BY total_balance DESC, u.user_id ASC
        """,
        (guild_id,),
    )


def create_admin_report(db: Database, guild_id: int, guild=None) -> Path:
    generated = _generated_at()
    guild_name = guild.name if guild is not None else f"Servidor {guild_id}"
    subtitle = (
        f"Servidor: {guild_name} | Guild ID: {guild_id} | "
        f"Generado: {generated.strftime('%Y-%m-%d %H:%M UTC')} | Acceso exclusivo para admins"
    )
    workbook, summary = _new_workbook("Reporte Integral Administrativo G3NESYS", subtitle)

    movements = db.fetch_all(
        "SELECT * FROM movements WHERE guild_id = ? ORDER BY id DESC",
        (guild_id,),
    )
    fines = db.fetch_all(
        "SELECT * FROM fines WHERE guild_id = ? ORDER BY id DESC",
        (guild_id,),
    )
    activities = db.fetch_all(
        """
        SELECT ac.*,
               COALESCE((SELECT SUM(slots) FROM activity_roles ar WHERE ar.activity_id = ac.id), 0) AS required_slots,
               COALESCE((SELECT COUNT(*) FROM activity_participants ap WHERE ap.activity_id = ac.id), 0) AS participants,
               COALESCE((SELECT COUNT(*) FROM asistencia_actividades aa WHERE aa.actividad_id = ac.id AND aa.estado = 'Confirmado'), 0) AS confirmed,
               COALESCE((SELECT COUNT(*) FROM asistencia_actividades aa WHERE aa.actividad_id = ac.id AND aa.estado = 'Ausente'), 0) AS absent
        FROM activities ac
        WHERE ac.guild_id = ?
        ORDER BY ac.id DESC
        """,
        (guild_id,),
    )
    payouts = db.fetch_all(
        """
        SELECT p.*,
               COALESCE((SELECT COUNT(*) FROM payout_participants pp WHERE pp.payout_id = p.id), 0) AS participants,
               COALESCE((SELECT SUM(amount) FROM payout_participants pp WHERE pp.payout_id = p.id), 0) AS assigned_amount
        FROM payouts p
        WHERE p.guild_id = ?
        ORDER BY p.id DESC
        """,
        (guild_id,),
    )
    withdrawals = db.fetch_all(
        "SELECT * FROM withdrawals WHERE guild_id = ? ORDER BY id DESC",
        (guild_id,),
    )
    user_stats = _fetch_user_stats(db, guild_id)
    rankings = caller_ranking(db, guild_id)
    rank_by_user = {
        int(row["user_id"]): (position, row)
        for position, row in enumerate(rankings, start=1)
    }

    treasury = db.fetch_one("SELECT balance FROM treasury WHERE guild_id = ?", (guild_id,))
    account_totals = db.fetch_one(
        """
        SELECT COALESCE(SUM(available), 0) AS available,
               COALESCE(SUM(retained), 0) AS retained,
               COALESCE(SUM(seized), 0) AS seized
        FROM accounts WHERE guild_id = ?
        """,
        (guild_id,),
    )
    pending_fines = db.fetch_one(
        """
        SELECT COUNT(*) AS total, COALESCE(SUM(amount), 0) AS amount
        FROM fines WHERE guild_id = ? AND status = 'Pendiente'
        """,
        (guild_id,),
    )
    active_caller_penalties = db.fetch_one(
        "SELECT COUNT(*) AS total FROM caller_penalties WHERE guild_id = ? AND active = 1",
        (guild_id,),
    )
    deposited = sum(
        int(row["distributable"])
        for row in payouts
        if row["status"] == "Depositos realizados"
    )
    movement_totals: dict[str, int] = defaultdict(int)
    for row in movements:
        movement_totals[str(row["type"])] += int(row["amount"])

    metrics = [
        ("Tesoreria", int(treasury["balance"]) if treasury else 0, True),
        ("Usuarios con registros", len(user_stats), False),
        ("Saldo disponible usuarios", int(account_totals["available"]), True),
        ("Saldo retenido usuarios", int(account_totals["retained"]), True),
        ("Saldo decomisado", int(account_totals["seized"]), True),
        ("Movimientos", len(movements), False),
        ("Multas pendientes", int(pending_fines["total"]), False),
        ("Monto multas pendientes", int(pending_fines["amount"]), True),
        ("Actividades", len(activities), False),
        ("Splits depositados", deposited, True),
        ("Callers autorizados", len(rankings), False),
        ("Callers penalizados", int(active_caller_penalties["total"]), False),
        ("Solicitudes de cobro", len(withdrawals), False),
        ("Pestañas del libro", 21, False),
    ]
    caller_chart_rows = [
        (_member_name(guild, int(row["user_id"])), int(row["score"]))
        for row in rankings
    ]
    _populate_summary(
        summary,
        metrics=metrics,
        caller_rows=caller_chart_rows,
        movement_rows=sorted(movement_totals.items(), key=lambda item: item[1], reverse=True),
    )

    user_rows = []
    for row in user_stats:
        user_id = int(row["user_id"])
        rank = rank_by_user.get(user_id)
        user_rows.append(
            [
                _id(user_id),
                _member_name(guild, user_id),
                int(row["available"]),
                int(row["retained"]),
                int(row["seized"]),
                int(row["total_balance"]),
                row["account_updated_at"],
                int(row["fines_total"]),
                int(row["fines_pending"]),
                int(row["fines_amount"]),
                int(row["activities_joined"]),
                int(row["attendance_confirmed"]),
                int(row["attendance_absent"]),
                int(row["attendance_justified"]),
                int(row["payouts_received"]),
                int(row["payouts_amount"]),
                int(row["withdrawals_total"]),
                int(row["withdrawals_requested"]),
                int(row["withdrawals_liquidated"]),
                int(row["transfers_sent"]),
                int(row["transfers_sent_amount"]),
                int(row["transfers_received"]),
                int(row["transfers_received_amount"]),
                "Si" if int(row["is_caller"]) else "No",
                rank[0] if rank else "",
                int(rank[1]["score"]) if rank else "",
                "Penalizado" if int(row["caller_penalized"]) else ("Activo" if rank else "No aplica"),
            ]
        )
    _add_detail_sheet(
        workbook,
        name="Usuarios",
        title="Usuarios y estadisticas consolidadas",
        subtitle=subtitle,
        headers=[
            "Usuario ID", "Nombre", "Disponible", "Retenido", "Decomisado", "Saldo total",
            "Cuenta actualizada", "Multas totales", "Multas pendientes", "Monto multas",
            "Actividades inscritas", "Asistencias", "Ausencias", "Justificadas",
            "Splits recibidos", "Plata recibida", "Cobros solicitados", "Monto solicitado",
            "Monto liquidado", "Transferencias enviadas", "Plata enviada",
            "Transferencias recibidas", "Plata recibida transf.", "Es caller", "Ranking caller",
            "Puntos caller", "Estado caller",
        ],
        rows=user_rows,
        currency_headers={
            "Disponible", "Retenido", "Decomisado", "Saldo total", "Monto multas",
            "Plata recibida", "Monto solicitado", "Monto liquidado", "Plata enviada",
            "Plata recibida transf.",
        },
        date_headers={"Cuenta actualizada"},
    )

    movement_rows = [
        [
            row["code"], row["type"], row["category"], _id(row["user_id"]),
            _member_name(guild, row["user_id"]), _id(row["counterparty_id"]),
            _member_name(guild, row["counterparty_id"]), int(row["amount"]),
            row["source_table"], row["source_id"], row["description"],
            _id(row["created_by"]), _member_name(guild, row["created_by"]), row["created_at"],
        ]
        for row in movements
    ]
    movement_headers = [
        "Codigo", "Tipo", "Categoria", "Usuario ID", "Usuario", "Contraparte ID",
        "Contraparte", "Monto", "Tabla origen", "ID origen", "Descripcion",
        "Creado por ID", "Creado por", "Fecha",
    ]
    _add_detail_sheet(
        workbook,
        name="Movimientos",
        title="Libro completo de movimientos",
        subtitle=subtitle,
        headers=movement_headers,
        rows=movement_rows,
        currency_headers={"Monto"},
        date_headers={"Fecha"},
    )
    transaction_rows = [row for row, source in zip(movement_rows, movements) if source["user_id"] or source["counterparty_id"]]
    _add_detail_sheet(
        workbook,
        name="Transacciones Usuarios",
        title="Transacciones relacionadas con usuarios",
        subtitle=subtitle,
        headers=movement_headers,
        rows=transaction_rows,
        currency_headers={"Monto"},
        date_headers={"Fecha"},
    )

    _add_detail_sheet(
        workbook,
        name="Multas",
        title="Registro integral de multas",
        subtitle=subtitle,
        headers=[
            "Codigo", "Usuario ID", "Usuario", "Monto", "Estado", "Motivo", "Origen",
            "Creada por ID", "Creada por", "Pagada por ID", "Metodo pago", "Movimiento ID",
            "Fecha creacion", "Fecha pago", "Cancelada por ID", "Fecha cancelacion",
            "Motivo cancelacion",
        ],
        rows=[
            [
                row["code"], _id(row["user_id"]), _member_name(guild, row["user_id"]),
                int(row["amount"]), row["status"], row["reason"], row["origin"],
                _id(row["created_by"]), _member_name(guild, row["created_by"]),
                _id(row["paid_by"]), row["payment_method"], row["movement_id"],
                row["created_at"], row["paid_at"], _id(row["cancelled_by"]),
                row["cancelled_at"], row["cancel_reason"],
            ]
            for row in fines
        ],
        currency_headers={"Monto"},
        date_headers={"Fecha creacion", "Fecha pago", "Fecha cancelacion"},
    )

    _add_detail_sheet(
        workbook,
        name="Actividades",
        title="Actividades y resultados",
        subtitle=subtitle,
        headers=[
            "Codigo", "Actividad", "Caller ID", "Caller", "Horario", "Estado",
            "Canal voz ID", "Cupos requeridos", "Participantes", "Confirmados", "Ausentes",
            "Plantilla ID", "Creada", "Check enviado", "Iniciada", "Finalizada", "Cancelada por ID",
            "Cancelacion justificada", "Motivo cancelacion", "Notas",
        ],
        rows=[
            [
                row["code"], row["name"], _id(row["caller_id"]),
                _member_name(guild, row["caller_id"]), row["horario"], row["status"],
                _id(row["voice_channel_id"]), int(row["required_slots"]),
                int(row["participants"]), int(row["confirmed"]), int(row["absent"]),
                row["template_id"], row["created_at"], row["check_sent_at"],
                row["started_at"], row["ended_at"],
                _id(row["cancelled_by"]),
                "Si" if row["cancellation_reputation_exempt"] else "No",
                row["cancellation_reason"], row["notes"],
            ]
            for row in activities
        ],
        date_headers={"Creada", "Check enviado", "Iniciada", "Finalizada"},
    )

    participants = db.fetch_all(
        """
        SELECT ac.code, ac.name AS activity_name, ap.user_id, ap.display_name,
               ar.name AS role_name, ar.emoji, ap.joined_at,
               aa.estado, aa.confirmo_boton, aa.confirmo_voz, aa.fecha_check,
               aa.genero_multa, aa.justificado_por, aa.observaciones,
               aa.voice_seconds, aa.participation_percent
        FROM activity_participants ap
        JOIN activities ac ON ac.id = ap.activity_id
        JOIN activity_roles ar ON ar.id = ap.role_id
        LEFT JOIN asistencia_actividades aa
            ON aa.actividad_id = ap.activity_id AND aa.usuario_id = ap.user_id
        WHERE ac.guild_id = ?
        ORDER BY ac.id DESC, ar.position ASC, ap.id ASC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Participantes",
        title="Participacion y asistencia por actividad",
        subtitle=subtitle,
        headers=[
            "Actividad ID", "Actividad", "Usuario ID", "Usuario actual", "Nombre registrado",
            "Rol arma", "Emoji", "Fecha registro", "Asistencia", "Confirmo boton",
            "Confirmo voz", "Fecha check", "Genero multa", "Justificado por ID", "Observaciones",
            "Segundos en voz", "Permanencia",
        ],
        rows=[
            [
                row["code"], row["activity_name"], _id(row["user_id"]),
                _member_name(guild, row["user_id"]), row["display_name"], row["role_name"],
                row["emoji"], row["joined_at"], row["estado"],
                "Si" if row["confirmo_boton"] else "No",
                "Si" if row["confirmo_voz"] else "No", row["fecha_check"],
                "Si" if row["genero_multa"] else "No", _id(row["justificado_por"]),
                row["observaciones"], int(row["voice_seconds"] or 0),
                float(row["participation_percent"] or 0),
            ]
            for row in participants
        ],
        percent_headers={"Permanencia"},
        date_headers={"Fecha registro", "Fecha check"},
    )

    join_requests = db.fetch_all(
        """
        SELECT jr.*, ac.code AS activity_code, ac.name AS activity_name
        FROM activity_join_requests jr
        JOIN activities ac ON ac.id = jr.activity_id
        WHERE jr.guild_id = ?
        ORDER BY jr.id DESC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Solicitudes Ingreso",
        title="Solicitudes tardias para unirse a actividades",
        subtitle=subtitle,
        headers=[
            "Registro ID", "Actividad ID", "Actividad", "Usuario ID", "Usuario",
            "Rol solicitado", "Estado", "Solicitado", "Revisado por ID", "Revisado",
        ],
        rows=[
            [
                row["id"], row["activity_code"], row["activity_name"], _id(row["user_id"]),
                _member_name(guild, row["user_id"]), row["requested_role"], row["status"],
                row["requested_at"], _id(row["reviewed_by"]), row["reviewed_at"],
            ]
            for row in join_requests
        ],
        date_headers={"Solicitado", "Revisado"},
    )

    voice_sessions = db.fetch_all(
        """
        SELECT vs.*, ac.code AS activity_code, ac.name AS activity_name
        FROM activity_voice_sessions vs
        JOIN activities ac ON ac.id = vs.activity_id
        WHERE vs.guild_id = ?
        ORDER BY vs.id DESC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Sesiones Voz",
        title="Intervalos de permanencia en canales de voz",
        subtitle=subtitle,
        headers=[
            "Sesion ID", "Actividad ID", "Actividad", "Usuario ID", "Usuario",
            "Entrada", "Salida", "Segundos",
        ],
        rows=[
            [
                row["id"], row["activity_code"], row["activity_name"], _id(row["user_id"]),
                _member_name(guild, row["user_id"]), row["joined_at"], row["left_at"],
                int(row["seconds"] or 0),
            ]
            for row in voice_sessions
        ],
        date_headers={"Entrada", "Salida"},
    )

    _add_detail_sheet(
        workbook,
        name="Splits",
        title="Splits y liquidacion de loot",
        subtitle=subtitle,
        headers=[
            "Codigo", "Actividad DB ID", "Caller ID", "Caller", "Estado", "Loot bruto",
            "Tasa mercado", "Reparaciones", "Otros gastos", "Porcentaje gremial",
            "Aporte gremial", "Repartible", "Participantes", "Monto asignado", "Notas",
            "Porcentaje caller", "Pago caller", "Creado", "Enviado admin", "Revisado por ID", "Revisado",
        ],
        rows=[
            [
                row["code"], row["activity_id"], _id(row["caller_id"]),
                _member_name(guild, row["caller_id"]), row["status"], int(row["gross_loot"]),
                float(row["market_rate_percent"]), int(row["repairs"]),
                int(row["other_expenses"]), float(row["guild_percent"]),
                int(row["guild_amount"]), int(row["distributable"]),
                int(row["participants"]), int(row["assigned_amount"]), row["notes"],
                float(row["caller_percent"] or 0), int(row["caller_amount"] or 0),
                row["created_at"], row["sent_to_admin_at"], _id(row["reviewed_by"]),
                row["reviewed_at"],
            ]
            for row in payouts
        ],
        currency_headers={
            "Loot bruto", "Reparaciones", "Otros gastos", "Aporte gremial", "Repartible",
            "Monto asignado",
            "Pago caller",
        },
        percent_headers={"Tasa mercado", "Porcentaje gremial", "Porcentaje caller"},
        date_headers={"Creado", "Enviado admin", "Revisado"},
    )

    payout_details = db.fetch_all(
        """
        SELECT p.code, p.status, pp.user_id, pp.participation_percent,
               pp.amount, pp.balance_type, pp.deposited_at
        FROM payout_participants pp
        JOIN payouts p ON p.id = pp.payout_id
        WHERE p.guild_id = ?
        ORDER BY p.id DESC, pp.id ASC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Detalle Splits",
        title="Detalle de participantes por Split",
        subtitle=subtitle,
        headers=[
            "Split", "Estado Split", "Usuario ID", "Usuario", "Participacion",
            "Monto", "Tipo saldo", "Fecha deposito",
        ],
        rows=[
            [
                row["code"], row["status"], _id(row["user_id"]),
                _member_name(guild, row["user_id"]), float(row["participation_percent"]),
                int(row["amount"]), row["balance_type"], row["deposited_at"],
            ]
            for row in payout_details
        ],
        currency_headers={"Monto"},
        percent_headers={"Participacion"},
        date_headers={"Fecha deposito"},
    )

    _add_detail_sheet(
        workbook,
        name="Cobros",
        title="Solicitudes y liquidaciones de cobro",
        subtitle=subtitle,
        headers=[
            "Codigo", "Usuario ID", "Usuario", "Monto solicitado", "Monto liquidado",
            "Estado", "Motivo", "Creado", "Aprobado por ID", "Aprobado",
            "Liquidado por ID", "Liquidado", "Rechazado por ID", "Rechazado",
            "Motivo rechazo",
        ],
        rows=[
            [
                row["code"], _id(row["user_id"]), _member_name(guild, row["user_id"]),
                int(row["amount_requested"]), int(row["amount_liquidated"] or 0),
                row["status"], row["reason"], row["created_at"], _id(row["approved_by"]),
                row["approved_at"], _id(row["liquidated_by"]), row["liquidated_at"],
                _id(row["rejected_by"]), row["rejected_at"], row["rejection_reason"],
            ]
            for row in withdrawals
        ],
        currency_headers={"Monto solicitado", "Monto liquidado"},
        date_headers={"Creado", "Aprobado", "Liquidado", "Rechazado"},
    )

    caller_meta = {
        int(row["user_id"]): row
        for row in db.fetch_all(
            "SELECT * FROM callers WHERE guild_id = ?",
            (guild_id,),
        )
    }
    _add_detail_sheet(
        workbook,
        name="Callers",
        title="Ranking y rendimiento de callers autorizados",
        subtitle=subtitle,
        headers=[
            "Ranking", "Caller ID", "Caller", "Puntos", "Estado", "Plata repartida",
            "Actividades creadas", "Completadas", "Canceladas", "Justificadas",
            "Asistencias", "Ausencias", "Agregado por ID", "Fecha autorizacion",
        ],
        rows=[
            [
                position, _id(row["user_id"]), _member_name(guild, row["user_id"]),
                int(row["score"]), "Penalizado" if int(row["penalized"]) else "Activo",
                int(row["distributed"]), int(row["activities_created"]),
                int(row["activities_completed"]), int(row["activities_cancelled"]),
                int(row["cancellations_exempt"]), int(row["attendances"]),
                int(row["absences"]), _id(caller_meta[int(row["user_id"])]["added_by"]),
                caller_meta[int(row["user_id"])]["created_at"],
            ]
            for position, row in enumerate(rankings, start=1)
        ],
        currency_headers={"Plata repartida"},
        date_headers={"Fecha autorizacion"},
    )

    caller_penalties = db.fetch_all(
        "SELECT * FROM caller_penalties WHERE guild_id = ? ORDER BY id DESC",
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Penaliz Callers",
        title="Historial de penalizaciones de callers",
        subtitle=subtitle,
        headers=[
            "Registro ID", "Caller ID", "Caller", "Puntos al penalizar", "Motivo",
            "Estado", "Penalizado", "Aviso DM", "Retirado por ID", "Retirado",
            "Rearmado",
        ],
        rows=[
            [
                row["id"], _id(row["user_id"]), _member_name(guild, row["user_id"]),
                int(row["score_at_penalty"]), row["reason"],
                "Activa" if int(row["active"]) else "Retirada", row["penalized_at"],
                row["notified_at"], _id(row["removed_by"]), row["removed_at"],
                "Si" if int(row["rearmed"]) else "No",
            ]
            for row in caller_penalties
        ],
        date_headers={"Penalizado", "Aviso DM", "Retirado"},
    )

    treasury_rows = [
        [movement_type, amount, sum(1 for row in movements if str(row["type"]) == movement_type)]
        for movement_type, amount in sorted(movement_totals.items())
    ]
    treasury_rows.insert(
        0,
        ["SALDO ACTUAL TESORERIA", int(treasury["balance"]) if treasury else 0, ""],
    )
    _add_detail_sheet(
        workbook,
        name="Tesoreria",
        title="Tesoreria y acumulados por tipo de movimiento",
        subtitle=subtitle,
        headers=["Concepto", "Monto", "Cantidad movimientos"],
        rows=treasury_rows,
        currency_headers={"Monto"},
    )

    templates = db.fetch_all(
        """
        SELECT t.*, COUNT(tr.id) AS roles,
               GROUP_CONCAT(
                   TRIM(COALESCE(tr.emoji || ' ', '') || tr.name || ' [' || tr.slots || ']'),
                   ' | '
               ) AS composition
        FROM templates t
        LEFT JOIN template_roles tr ON tr.template_id = t.id
        WHERE t.guild_id = ?
        GROUP BY t.id
        ORDER BY t.id DESC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Plantillas",
        title="Plantillas de actividades",
        subtitle=subtitle,
        headers=[
            "Plantilla DB ID", "Nombre", "Actividad base", "Horario base", "Descripcion",
            "Cantidad roles", "Composicion", "Creada por ID", "Creada por", "Fecha",
        ],
        rows=[
            [
                row["id"], row["name"], row["activity_name"], row["default_time"],
                row["description"], int(row["roles"]), row["composition"],
                _id(row["created_by"]), _member_name(guild, row["created_by"]), row["created_at"],
            ]
            for row in templates
        ],
        date_headers={"Fecha"},
    )

    custom_settings = {
        str(row["key"]): row
        for row in db.fetch_all(
            "SELECT * FROM guild_settings WHERE guild_id = ?",
            (guild_id,),
        )
    }
    setting_keys = sorted(set(DEFAULT_SETTINGS) | set(custom_settings))
    _add_detail_sheet(
        workbook,
        name="Configuracion",
        title="Configuracion efectiva del servidor",
        subtitle=subtitle,
        headers=["Clave", "Valor efectivo", "Origen", "Actualizado"],
        rows=[
            [
                key,
                db.get_setting(guild_id, key),
                "Personalizada" if key in custom_settings else "Predeterminada",
                custom_settings[key]["updated_at"] if key in custom_settings else "",
            ]
            for key in setting_keys
        ],
        date_headers={"Actualizado"},
    )

    activity_penalties = db.fetch_all(
        """
        SELECT * FROM penalizacion_actividades
        WHERE guild_id = ? ORDER BY id DESC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Penaliz Actividades",
        title="Penalizaciones generales de actividades",
        subtitle=subtitle,
        headers=[
            "Registro ID", "Usuario ID", "Usuario", "Motivo", "Origen", "Estado",
            "Fecha ingreso", "Removido por ID", "Fecha remocion", "Observaciones",
        ],
        rows=[
            [
                row["id"], _id(row["usuario_id"]), _member_name(guild, row["usuario_id"]),
                row["motivo"], row["origen"], "Activa" if int(row["activo"]) else "Retirada",
                row["fecha_ingreso"], _id(row["removido_por"]), row["fecha_remocion"],
                row["observaciones"],
            ]
            for row in activity_penalties
        ],
        date_headers={"Fecha ingreso", "Fecha remocion"},
    )

    audit_rows = db.fetch_all(
        "SELECT * FROM audit_logs WHERE guild_id = ? ORDER BY id DESC",
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Auditoria",
        title="Bitacora administrativa y automatica",
        subtitle=subtitle,
        headers=[
            "Registro ID", "Admin actor ID", "Admin actor", "Accion", "Usuario afectado ID",
            "Usuario afectado", "Monto", "Sistema", "Observacion", "Fecha",
        ],
        rows=[
            [
                row["id"], _id(row["admin_id"]), _member_name(guild, row["admin_id"]),
                row["action"], _id(row["affected_user_id"]),
                _member_name(guild, row["affected_user_id"]), int(row["amount"] or 0),
                row["system"], row["observation"], row["created_at"],
            ]
            for row in audit_rows
        ],
        currency_headers={"Monto"},
        date_headers={"Fecha"},
    )

    dm_rows = db.fetch_all(
        """
        SELECT * FROM dm_logs
        WHERE guild_id = ?
        ORDER BY id DESC
        """,
        (guild_id,),
    )
    _add_detail_sheet(
        workbook,
        name="Notificaciones",
        title="Entrega de mensajes directos",
        subtitle=subtitle,
        headers=["Registro ID", "Usuario ID", "Usuario", "Accion", "Resultado", "Error", "Fecha"],
        rows=[
            [
                row["id"], _id(row["user_id"]), _member_name(guild, row["user_id"]),
                row["action"], "Entregado" if int(row["success"]) else "Fallido",
                row["error"], row["created_at"],
            ]
            for row in dm_rows
        ],
        date_headers={"Fecha"},
    )

    validation_rows = []
    for payout in payouts:
        difference = int(payout["assigned_amount"]) - int(payout["distributable"])
        validation_rows.append(
            [
                f"Split {payout['code']}", int(payout["assigned_amount"]),
                int(payout["distributable"]), difference,
                "OK" if difference == 0 else "REVISAR",
                "La suma de participantes debe coincidir con el monto repartible.",
            ]
        )
        net_amount = (
            int(payout["gross_loot"])
            - int(round(int(payout["gross_loot"]) * (float(payout["market_rate_percent"]) / 100)))
            - int(payout["repairs"])
            - int(payout["other_expenses"])
        )
        allocated = (
            int(payout["guild_amount"])
            + int(payout["caller_amount"] or 0)
            + int(payout["distributable"])
        )
        allocation_difference = allocated - net_amount
        validation_rows.append(
            [
                f"Estructura Split {payout['code']}", allocated, net_amount,
                allocation_difference, "OK" if allocation_difference == 0 else "REVISAR",
                "Gremio + caller + participantes debe coincidir con el neto del loot.",
            ]
        )
    for table in ("activities", "fines", "withdrawals", "payouts", "movements"):
        check = db.fetch_one(
            f"SELECT COUNT(*) AS total, COUNT(DISTINCT code) AS distinct_codes FROM {table} WHERE guild_id = ?",
            (guild_id,),
        )
        difference = int(check["total"]) - int(check["distinct_codes"])
        validation_rows.append(
            [
                f"Codigos unicos: {table}", int(check["total"]),
                int(check["distinct_codes"]), difference,
                "OK" if difference == 0 else "REVISAR",
                "No debe haber codigos repetidos dentro del mismo servidor.",
            ]
        )
    treasury_balance = int(treasury["balance"]) if treasury else 0
    validation_rows.append(
        [
            "Tesoreria no negativa", treasury_balance, 0, treasury_balance,
            "OK" if treasury_balance >= 0 else "REVISAR",
            "La tesoreria no debe quedar por debajo de cero.",
        ]
    )
    checks = _add_detail_sheet(
        workbook,
        name="Validaciones",
        title="Controles de integridad del reporte",
        subtitle=subtitle,
        headers=["Control", "Actual", "Esperado", "Diferencia", "Estado", "Notas"],
        rows=validation_rows,
        currency_headers={"Actual", "Esperado", "Diferencia"},
    )
    for row_number in range(5, 5 + len(validation_rows)):
        state = checks.cell(row_number, 5)
        state.fill = PatternFill(
            "solid",
            fgColor="D9EAD3" if state.value == "OK" else "F4CCCC",
        )
        state.font = Font(bold=True, color="274E13" if state.value == "OK" else "990000")

    timestamp = generated.strftime("%Y%m%d-%H%M%S")
    return _save_workbook(
        workbook,
        f"reporte-admin-g3nesys-{guild_id}-{timestamp}.xlsx",
    )


def create_caller_report(db: Database, guild_id: int, user_id: int, guild=None) -> Path:
    generated = _generated_at()
    guild_name = guild.name if guild is not None else f"Servidor {guild_id}"
    caller_name = _member_name(guild, user_id)
    rankings = caller_ranking(db, guild_id)
    caller_row = None
    caller_position = None
    for position, row in enumerate(rankings, start=1):
        if int(row["user_id"]) == user_id:
            caller_row = row
            caller_position = position
            break
    if caller_row is None:
        raise ValueError("No tienes un perfil de caller autorizado en este servidor.")

    subtitle = (
        f"Caller: {caller_name} ({user_id}) | Servidor: {guild_name} | "
        f"Generado: {generated.strftime('%Y-%m-%d %H:%M UTC')} | Reporte personal"
    )
    workbook, summary = _new_workbook("Mi Reporte de Caller G3NESYS", subtitle)
    penalties = db.fetch_all(
        """
        SELECT * FROM caller_penalties
        WHERE guild_id = ? AND user_id = ?
        ORDER BY id DESC
        """,
        (guild_id, user_id),
    )
    activities = db.fetch_all(
        """
        SELECT ac.*,
               COALESCE((SELECT SUM(slots) FROM activity_roles ar WHERE ar.activity_id = ac.id), 0) AS required_slots,
               COALESCE((SELECT COUNT(*) FROM activity_participants ap WHERE ap.activity_id = ac.id), 0) AS participants,
               COALESCE((SELECT COUNT(*) FROM asistencia_actividades aa WHERE aa.actividad_id = ac.id AND aa.estado = 'Confirmado'), 0) AS confirmed,
               COALESCE((SELECT COUNT(*) FROM asistencia_actividades aa WHERE aa.actividad_id = ac.id AND aa.estado = 'Ausente'), 0) AS absent
        FROM activities ac
        WHERE ac.guild_id = ? AND ac.caller_id = ?
        ORDER BY ac.id DESC
        """,
        (guild_id, user_id),
    )
    attendance = db.fetch_all(
        """
        SELECT ac.code, ac.name, ac.horario, ac.status AS activity_status,
               aa.estado, aa.confirmo_boton, aa.confirmo_voz, aa.fecha_check,
               aa.voice_seconds, aa.participation_percent, ar.name AS role_name
        FROM asistencia_actividades aa
        JOIN activities ac ON ac.id = aa.actividad_id
        LEFT JOIN activity_participants ap
            ON ap.activity_id = aa.actividad_id AND ap.user_id = aa.usuario_id
        LEFT JOIN activity_roles ar ON ar.id = ap.role_id
        WHERE ac.guild_id = ? AND aa.usuario_id = ?
        ORDER BY ac.id DESC
        """,
        (guild_id, user_id),
    )
    payouts = db.fetch_all(
        """
        SELECT p.*,
               COALESCE((SELECT COUNT(*) FROM payout_participants pp WHERE pp.payout_id = p.id), 0) AS participants
        FROM payouts p
        WHERE p.guild_id = ? AND p.caller_id = ?
        ORDER BY p.id DESC
        """,
        (guild_id, user_id),
    )
    templates = db.fetch_all(
        """
        SELECT t.*, COUNT(tr.id) AS roles,
               GROUP_CONCAT(
                   TRIM(COALESCE(tr.emoji || ' ', '') || tr.name || ' [' || tr.slots || ']'),
                   ' | '
               ) AS composition
        FROM templates t
        LEFT JOIN template_roles tr ON tr.template_id = t.id
        WHERE t.guild_id = ? AND t.created_by = ?
        GROUP BY t.id ORDER BY t.id DESC
        """,
        (guild_id, user_id),
    )

    status_counts: dict[str, int] = defaultdict(int)
    for activity in activities:
        status_counts[str(activity["status"])] += 1
    metrics = [
        ("Ranking", caller_position, False),
        ("Puntos", int(caller_row["score"]), False),
        ("Estado", "Penalizado" if int(caller_row["penalized"]) else "Activo", False),
        ("Plata repartida", int(caller_row["distributed"]), True),
        ("Actividades creadas", int(caller_row["activities_created"]), False),
        ("Completadas", int(caller_row["activities_completed"]), False),
        ("Canceladas", int(caller_row["activities_cancelled"]), False),
        ("Cancelaciones justificadas", int(caller_row["cancellations_exempt"]), False),
        ("Asistencias", int(caller_row["attendances"]), False),
        ("Ausencias", int(caller_row["absences"]), False),
        ("Penalizaciones historicas", len(penalties), False),
        ("Plantillas creadas", len(templates), False),
    ]
    _populate_summary(
        summary,
        metrics=metrics,
        caller_rows=[(caller_name, int(caller_row["score"]))],
        movement_rows=sorted(status_counts.items(), key=lambda item: item[1], reverse=True),
        caller_chart_title="Puntuacion actual",
        movement_chart_title="Actividades por estado",
    )

    _add_detail_sheet(
        workbook,
        name="Mis Actividades",
        title="Actividades dirigidas",
        subtitle=subtitle,
        headers=[
            "Codigo", "Actividad", "Horario", "Estado", "Cupos requeridos", "Participantes",
            "Confirmados", "Ausentes", "Creada", "Check enviado", "Iniciada", "Finalizada",
            "Cancelacion justificada", "Motivo cancelacion", "Notas",
        ],
        rows=[
            [
                row["code"], row["name"], row["horario"], row["status"],
                int(row["required_slots"]), int(row["participants"]), int(row["confirmed"]),
                int(row["absent"]), row["created_at"], row["check_sent_at"],
                row["started_at"], row["ended_at"],
                "Si" if row["cancellation_reputation_exempt"] else "No",
                row["cancellation_reason"], row["notes"],
            ]
            for row in activities
        ],
        date_headers={"Creada", "Check enviado", "Iniciada", "Finalizada"},
    )
    _add_detail_sheet(
        workbook,
        name="Mi Asistencia",
        title="Historial personal de asistencia",
        subtitle=subtitle,
        headers=[
            "Actividad ID", "Actividad", "Horario", "Estado actividad", "Rol arma",
            "Asistencia", "Confirmo boton", "Confirmo voz", "Fecha check",
            "Segundos en voz", "Permanencia",
        ],
        rows=[
            [
                row["code"], row["name"], row["horario"], row["activity_status"],
                row["role_name"] or "Caller sin cupo de composicion", row["estado"],
                "Si" if row["confirmo_boton"] else "No",
                "Si" if row["confirmo_voz"] else "No", row["fecha_check"],
                int(row["voice_seconds"] or 0), float(row["participation_percent"] or 0),
            ]
            for row in attendance
        ],
        percent_headers={"Permanencia"},
        date_headers={"Fecha check"},
    )
    _add_detail_sheet(
        workbook,
        name="Mis Splits",
        title="Splits generados como caller",
        subtitle=subtitle,
        headers=[
            "Codigo", "Actividad DB ID", "Estado", "Loot bruto", "Tasa mercado",
            "Reparaciones", "Otros gastos", "Porcentaje gremial", "Aporte gremial",
            "Porcentaje caller", "Pago caller", "Repartible", "Participantes", "Creado", "Enviado admin", "Revisado",
        ],
        rows=[
            [
                row["code"], row["activity_id"], row["status"], int(row["gross_loot"]),
                float(row["market_rate_percent"]), int(row["repairs"]),
                int(row["other_expenses"]), float(row["guild_percent"]),
                int(row["guild_amount"]), float(row["caller_percent"] or 0),
                int(row["caller_amount"] or 0), int(row["distributable"]),
                int(row["participants"]), row["created_at"], row["sent_to_admin_at"],
                row["reviewed_at"],
            ]
            for row in payouts
        ],
        currency_headers={
            "Loot bruto", "Reparaciones", "Otros gastos", "Aporte gremial", "Pago caller", "Repartible",
        },
        percent_headers={"Tasa mercado", "Porcentaje gremial", "Porcentaje caller"},
        date_headers={"Creado", "Enviado admin", "Revisado"},
    )
    _add_detail_sheet(
        workbook,
        name="Mis Penalizaciones",
        title="Historial personal de penalizaciones de caller",
        subtitle=subtitle,
        headers=[
            "Registro ID", "Puntos al penalizar", "Motivo", "Estado", "Penalizado",
            "Aviso DM", "Retirado por ID", "Retirado", "Rearmado",
        ],
        rows=[
            [
                row["id"], int(row["score_at_penalty"]), row["reason"],
                "Activa" if int(row["active"]) else "Retirada", row["penalized_at"],
                row["notified_at"], _id(row["removed_by"]), row["removed_at"],
                "Si" if int(row["rearmed"]) else "No",
            ]
            for row in penalties
        ],
        date_headers={"Penalizado", "Aviso DM", "Retirado"},
    )
    _add_detail_sheet(
        workbook,
        name="Mis Plantillas",
        title="Plantillas creadas por el caller",
        subtitle=subtitle,
        headers=[
            "Plantilla DB ID", "Nombre", "Actividad base", "Horario base", "Descripcion",
            "Cantidad roles", "Composicion", "Fecha",
        ],
        rows=[
            [
                row["id"], row["name"], row["activity_name"], row["default_time"],
                row["description"], int(row["roles"]), row["composition"], row["created_at"],
            ]
            for row in templates
        ],
        date_headers={"Fecha"},
    )

    timestamp = generated.strftime("%Y%m%d-%H%M%S")
    return _save_workbook(
        workbook,
        f"reporte-caller-{guild_id}-{user_id}-{timestamp}.xlsx",
    )
